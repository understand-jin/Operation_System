import io
import numpy as np
import pandas as pd
from openpyxl import load_workbook


def apply_template(df: pd.DataFrame, template_bytes: bytes) -> bytes:
    """
    템플릿 xlsx에 df 데이터를 삽입하여 스타일 적용된 바이트 반환.
    - 템플릿 1행에 이미 헤더가 있으면 → 2행부터 데이터 기록
    - 1행이 비어 있으면 → 1행에 헤더, 2행부터 데이터 기록
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    # 템플릿 1행에 기존 헤더가 있는지 확인
    first_row_vals = [ws.cell(row=1, column=c).value for c in range(1, len(df.columns) + 1)]
    has_headers = any(v is not None for v in first_row_vals)

    if not has_headers:
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

    data_start = 2
    for row_idx, row_vals in enumerate(df.values, data_start):
        for col_idx, value in enumerate(row_vals, 1):
            # pandas NA / float nan → None으로 변환
            if value is pd.NA or (isinstance(value, float) and np.isnan(value)):
                value = None
            ws.cell(row=row_idx, column=col_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
